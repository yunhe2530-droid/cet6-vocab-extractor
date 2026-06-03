"""六级别黄标单词提取引擎"""

import io
import re
import time
from collections import defaultdict

import fitz
import openpyxl
import pytesseract
import requests
from openpyxl.styles import Font, Alignment, PatternFill
from PIL import Image

from cet6_phrases import CET6_PHRASES, check_phrase

# ── 配置 ──────────────────────────────────────────
YELLOW = (0.973, 0.890, 0.518)  # PDF 黄标 RGB float
YELLOW_TOLERANCE = 0.001
TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

# 停用词表（基础功能词，不应收入单词本）
STOP_WORDS = {
    "the", "a", "an", "and", "or", "but", "if", "because", "so", "than",
    "that", "this", "these", "those", "what", "which", "who", "whom", "whose",
    "when", "where", "why", "how", "all", "each", "every", "both", "few",
    "more", "most", "some", "any", "no", "not", "only", "own", "same",
    "as", "at", "by", "for", "in", "of", "on", "to", "from", "with",
    "into", "through", "during", "before", "after", "above", "below",
    "between", "under", "over", "up", "down", "out", "off", "about",
    "against", "toward", "upon", "without", "within",
    "be", "is", "are", "was", "were", "been", "being", "am",
    "have", "has", "had", "having",
    "do", "does", "did", "doing",
    "will", "would", "shall", "should", "may", "might", "can", "could",
    "must", "need", "dare",
    "i", "you", "he", "she", "it", "we", "they", "me", "him", "her", "us", "them",
    "my", "your", "his", "its", "our", "their", "mine", "yours", "hers",
    "itself", "himself", "herself", "myself", "yourself", "ourselves", "themselves",
    "this", "that", "these", "those",
    "who", "whom", "whose", "which", "what", "whoever", "whatever", "whichever",
    "someone", "somebody", "something", "anyone", "anybody", "anything",
    "everyone", "everybody", "everything", "nobody", "nothing", "noone",
    "there", "here", "where", "when", "why", "how",
    "very", "just", "quite", "too", "much", "little", "enough",
    "such", "so", "also", "even", "still", "already", "yet",
    "now", "then", "often", "always", "never", "sometimes", "usually",
    "well", "however", "therefore", "thus", "hence",
    "indeed", "perhaps", "maybe", "otherwise",
    "get", "got", "gets", "getting",
    "say", "said", "says", "telling", "told", "tell", "tells",
    "make", "made", "makes", "making",
    "go", "went", "gone", "goes", "going",
    "come", "came", "comes", "coming",
    "take", "took", "taken", "takes", "taking",
    "see", "saw", "seen", "sees", "seeing",
    "know", "knew", "known", "knows", "knowing",
    "think", "thought", "thinks", "thinking",
    "give", "gave", "given", "gives", "giving",
    "find", "found", "finds", "finding",
    "want", "wanted", "wants", "wanting",
    "like", "liked", "likes", "liking",
    "use", "used", "uses", "using",
    "look", "looked", "looks", "looking",
    "work", "worked", "works", "working",
    "seem", "seemed", "seems", "seeming",
    "try", "tried", "tries", "trying",
    "ask", "asked", "asks", "asking",
    "need", "needed", "needs", "needing",
    "feel", "felt", "feels", "feeling",
    "help", "helped", "helps", "helping",
    "keep", "kept", "keeps", "keeping",
    "put", "puts", "putting",
    "set", "sets", "setting",
    "let", "lets", "letting",
    "mean", "meant", "means", "meaning",
    "begin", "began", "begun", "begins", "beginning",
    "happen", "happened", "happens", "happening",
    "show", "showed", "shown", "shows", "showing",
    "bring", "brought", "brings", "bringing",
    "turn", "turned", "turns", "turning",
    "call", "called", "calls", "calling",
    "provide", "provided", "provides", "providing",
    "consider", "considered", "considers", "considering",
    "appear", "appeared", "appears", "appearing",
    "expect", "expected", "expects", "expecting",
    "include", "included", "includes", "including",
    "change", "changed", "changes", "changing",
    "lead", "led", "leads", "leading",
    "learn", "learned", "learns", "learning",
    "live", "lived", "lives", "living",
    "believe", "believed", "believes", "believing",
    "hold", "held", "holds", "holding",
    "write", "wrote", "written", "writes", "writing",
    "stand", "stood", "stands", "standing",
    "actually", "almost", "another", "around", "away", "back",
    "because", "become", "becomes", "becoming", "became",
    "being", "best", "better", "big", "bigger", "larger",
    "cannot", "cause", "causes", "certain", "certainly",
    "clear", "clearly", "close", "come", "course", "different",
    "each", "early", "else", "end", "enough", "especially",
    "example", "fact", "far", "finally", "first", "following",
    "forward", "further", "general", "generally", "getting",
    "given", "going", "good", "great", "group", "having",
    "help", "high", "highly", "idea", "important", "instead",
    "interest", "interested", "interesting", "kind",
    "large", "last", "later", "least", "leave", "leaving",
    "less", "left", "long", "longer", "made", "making",
    "man", "men", "might", "money", "moreover", "most",
    "mostly", "moving", "much", "name", "namely",
    "necessary", "next", "nonetheless", "nor",
    "nothing", "notice", "number", "obtain", "obtained",
    "obvious", "obviously", "often", "old", "once", "one",
    "ones", "order", "other", "others", "otherwise",
    "particular", "particularly", "partly", "people",
    "per", "perhaps", "person", "place", "point",
    "possible", "presumably", "previous", "previously",
    "primarily", "probably", "proper", "quite",
    "rather", "really", "reason", "reasonably", "recent",
    "recently", "regard", "regarding", "regardless",
    "related", "relatively", "respect", "respectively",
    "result", "resulting", "right", "second", "serious",
    "seriously", "several", "short", "shortly", "significantly",
    "similar", "similarly", "simply", "since", "slightly",
    "someone", "somewhat", "specially", "specific",
    "specifically", "state", "states", "still", "strong",
    "strongly", "subject", "subsequently", "substantial",
    "substantially", "successful", "successfully", "sufficient",
    "sufficiently", "sure", "surely", "taking", "third",
    "thorough", "thoroughly", "throughout", "together",
    "toward", "turned", "turning", "turns", "typical",
    "typically", "unless", "unlikely", "upon", "useful",
    "various", "vary", "varied", "varies", "varying",
    "way", "ways", "whereas", "whether", "whole", "wide",
    "widely", "willing", "within", "wonder", "wondering",
    "worth", "yet", "young", "yourself", "yourselves",
    "it's", "its", "many", "past", "two", "years", "words",
    "life", "hope", "way", "ways", "day", "days", "time",
    "thing", "things", "people", "person", "place", "part",
    "world", "year", "work", "hand", "hands", "head",
    "eye", "eyes", "face", "mind", "heart",
    "long", "short", "big", "small", "old", "new", "good",
    "great", "high", "low", "large", "full",
    "come", "came", "go", "went", "gone", "take", "took",
    "give", "gave", "put", "set", "let",
    "say", "said", "tell", "told", "ask", "asked",
    "another", "many", "much", "lots", "lots of",
    "always", "never", "often", "usually",
    "today", "yesterday", "tomorrow", "now", "then",
    "first", "second", "third", "last", "next",
    "may", "can", "could", "would", "should", "might",
    "must", "shall", "will", "need", "dare", "ought",
    "yes", "no", "sure", "ok", "okay", "well",
    "please", "thanks", "thank",
    "maybe", "perhaps", "probably", "possibly",
    "actually", "basically", "generally", "literally",
    "essentially", "ultimately", "eventually",
}

# ── 颜色检测 ──────────────────────────────────────

def is_yellow(fill):
    """判断 fill 是否匹配黄标颜色"""
    if fill is None:
        return False
    try:
        return (abs(fill[0] - YELLOW[0]) < YELLOW_TOLERANCE and
                abs(fill[1] - YELLOW[1]) < YELLOW_TOLERANCE and
                abs(fill[2] - YELLOW[2]) < YELLOW_TOLERANCE)
    except (TypeError, IndexError):
        return False


# ── 原生 PDF 提取 ──────────────────────────────────

def extract_native(pdf_bytes, filename=""):
    """从原生文本 PDF 提取黄标区域内的单词"""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page_words = {}

    for page_num, page in enumerate(doc):
        # 收集该页所有黄色 rect
        yellow_rects = []
        for drawing in page.get_drawings():
            if drawing.get("fill") and is_yellow(drawing["fill"]):
                yellow_rects.append(drawing["rect"])
            # 也检测 fill 在颜色属性里的情况
            for item in drawing.get("items", []):
                if len(item) >= 3:
                    fill = item[2] if isinstance(item[2], (tuple, list)) else None
                    if fill and is_yellow(fill) and len(item) > 1:
                        try:
                            yellow_rects.append(item[1])
                        except Exception:
                            pass

        # 合并重叠 rect
        yellow_rects = _merge_rects(yellow_rects)
        if not yellow_rects:
            continue

        # 提取页内文字
        blocks = page.get_text("dict")["blocks"]
        page_words[page_num] = []

        for block in blocks:
            if block.get("type") != 0:  # text block only
                continue
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    bbox = fitz.Rect(span["bbox"])
                    text = span["text"].strip()
                    if not text:
                        continue
                    # 检测是否落在黄色 rect 内
                    for yr in yellow_rects:
                        overlap = bbox.intersect(yr)
                        if overlap and overlap.get_area() > 0:
                            page_words[page_num].append({
                                "text": text,
                                "bbox": bbox,
                                "font": span.get("font", ""),
                                "size": span.get("size", 0),
                            })
                            break

    doc.close()
    return page_words


def _merge_rects(rects):
    """合并重叠或相邻的矩形"""
    if not rects:
        return []
    merged = []
    sorted_rects = sorted(rects, key=lambda r: (r.y0, r.x0))
    current = sorted_rects[0]
    for r in sorted_rects[1:]:
        if (abs(r.y0 - current.y0) < 5 and abs(r.y1 - current.y1) < 5
                and r.x0 <= current.x1 + 5):
            current = fitz.Rect(current.x0, current.y0,
                                max(current.x1, r.x1), max(current.y1, r.y1))
        else:
            merged.append(current)
            current = r
    merged.append(current)
    return merged


# ── 扫描件 PDF 提取 ───────────────────────────────

def extract_scanned(pdf_bytes, filename=""):
    """从扫描件 PDF 提取黄标区域→OCR"""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page_words = {}

    for page_num, page in enumerate(doc):
        # 获取黄色 rect
        yellow_rects = _get_yellow_rects(page)
        if not yellow_rects:
            continue

        # 获取底图（不含 overlay）
        images = page.get_images(full=True)
        if not images:
            continue

        # 取分辨率最高的底图
        best_img = max(images, key=lambda x: (x[2], x[3]))
        xref = best_img[0]
        pix = fitz.Pixmap(doc, xref)
        if pix.n > 4:
            pix = fitz.Pixmap(fitz.csRGB, pix)

        img_bytes = pix.tobytes("png")
        pil_img = Image.open(io.BytesIO(img_bytes))
        img_w, img_h = pil_img.size

        # 获取图片在 PDF 页面的放置位置
        img_info_list = page.get_image_info()
        img_bbox = None
        for info in img_info_list:
            if info.get("xref") == xref or True:  # 用第一个图片的 bbox
                img_bbox = fitz.Rect(info["bbox"])
                break
        if not img_bbox:
            # fallback: 用页面尺寸
            img_bbox = page.rect

        # 图片在 PDF 空间中的实际尺寸
        pdf_img_w = img_bbox.width
        pdf_img_h = img_bbox.height

        # 对每个黄色 rect 进行 OCR
        page_words[page_num] = []
        for yr in yellow_rects:
            # 只处理与图片区域重叠的黄标
            overlap = yr.intersect(img_bbox)
            if overlap.get_area() <= 0:
                continue

            # 坐标映射：PDF 空间 → 图片像素空间
            scale_x = img_w / pdf_img_w
            scale_y = img_h / pdf_img_h
            x0 = int((yr.x0 - img_bbox.x0) * scale_x)
            y0 = int((yr.y0 - img_bbox.y0) * scale_y)
            x1 = int((yr.x1 - img_bbox.x0) * scale_x)
            y1 = int((yr.y1 - img_bbox.y0) * scale_y)
            y0 = int(yr.y0 * scale_y)
            x1 = int(yr.x1 * scale_x)
            y1 = int(yr.y1 * scale_y)
            x0, y0, x1, y1 = max(0, x0), max(0, y0), min(img_w, x1), min(img_h, y1)
            if x1 <= x0 or y1 <= y0:
                continue

            # 加裁边 5px 提高识别率
            pad = 5
            crop = pil_img.crop((max(0, x0-pad), max(0, y0-pad),
                                 min(img_w, x1+pad), min(img_h, y1+pad)))
            # 放大提高 OCR 准确率
            crop = crop.resize((crop.width * 3, crop.height * 3), Image.LANCZOS)

            try:
                text = pytesseract.image_to_string(crop, lang="eng",
                                                   config="--psm 7 --oem 3")
                text = text.strip()
                if text:
                    page_words[page_num].append({
                        "text": text,
                        "bbox": yr,
                    })
            except Exception:
                continue

    doc.close()
    return page_words


def _get_yellow_rects(page):
    """从 page 提取所有黄色 rect"""
    rects = []
    for drawing in page.get_drawings():
        if drawing.get("fill") and is_yellow(drawing["fill"]):
            rects.append(drawing["rect"])
        for item in drawing.get("items", []):
            if len(item) >= 3:
                fill = item[2] if isinstance(item[2], (tuple, list)) else None
                if fill and is_yellow(fill) and len(item) > 1:
                    try:
                        rects.append(item[1])
                    except Exception:
                        pass
    return _merge_rects(rects)


# ── 单词净化 ──────────────────────────────────────

def clean_text(text):
    """清洗 OCR/提取的文字：去标点、拆分单词"""
    # 统一 smart quotes
    text = text.replace('‘', "'").replace('’', "'")
    text = text.replace('“', '"').replace('”', '"')
    # 移除冗余字符
    text = re.sub(r'[–—·•●○□■▲△▼▽◆◇★☆※→←↑↓♯♭♪♫]', ' ', text)
    # 按非字母/连字符/撇号分割
    words = re.findall(r"[A-Za-z]+(?:['-][A-Za-z]+)*", text)
    return [w for w in words if len(w) > 1]


def is_valid_word(w):
    """过滤无效单词和停用词"""
    if len(w) <= 2:
        return False
    if len(w) > 30:
        return False
    if w.isupper() and len(w) > 8:
        return False
    if w.lower() in STOP_WORDS:
        return False
    # 过滤明显 OCR 噪点：内部有大写字母混乱（如 "yYRemove"）
    if re.search(r'[a-z][A-Z][a-z]', w):
        return False
    # 过滤连续重复字符 4+（如 "succeec"、"aaaab"）
    if re.search(r'(.)\1{3,}', w):
        return False
    # 只保留字母、连字符、撇号
    if not re.match(r"^[A-Za-z'\-]+$", w):
        return False
    return True


def find_phrases(words):
    """从单词列表中检测已知短语"""
    found = []
    i = 0
    while i < len(words):
        # 尝试从 i 开始匹配最长短语
        matched = False
        for n in range(min(6, len(words) - i), 1, -1):  # 最长6词
            candidate = ' '.join(words[i:i+n]).lower()
            if candidate in CET6_PHRASES:
                found.append(candidate)
                i += n
                matched = True
                break
        if not matched:
            i += 1
    return found


# ── 词典查询 ──────────────────────────────────────

def query_youdao(word, retries=3):
    """查询 Youdao 词典，返回 (中文释义, 词性)"""
    url = f"https://dict.youdao.com/w/eng/{word}"
    headers = {
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/120.0.0.0 Safari/537.36"),
        "Accept": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    }

    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=headers, timeout=10)
            if resp.status_code != 200:
                continue

            html = resp.text

            # 优先查词性+释义
            pos_meaning = []
            # 模式1: <span class="pos">n.</span><span class="trans">释义</span>
            pos_pattern = re.findall(
                r'<span\s+class="pos"[^>]*>(.*?)</span>\s*'
                r'<span\s+class="trans"[^>]*>(.*?)</span>',
                html, re.DOTALL
            )
            for pos, trans in pos_pattern:
                pos = pos.strip()
                trans = trans.strip().replace("\n", "")
                trans = re.sub(r'<[^>]+>', '', trans)
                pos_meaning.append(f"{pos} {trans}")

            # 模式2: <li><span class="word">def</span></li>
            li_pattern = re.findall(
                r'<li[^>]*>\s*<span[^>]*class="[^"]*trans[^"]*"[^>]*>'
                r'(.*?)</span>',
                html, re.DOTALL
            )
            for m in li_pattern:
                m = re.sub(r'<[^>]+>', '', m).strip()
                if m and '[' not in m and len(m) > 1:
                    if m not in pos_meaning:
                        pos_meaning.append(m)

            # 模式3: <div class="def">释义</div>
            if not pos_meaning:
                def_pattern = re.findall(
                    r'<div\s+class="def"[^>]*>(.*?)</div>',
                    html, re.DOTALL
                )
                for d in def_pattern:
                    d = re.sub(r'<[^>]+>', '', d).strip()
                    if d and len(d) > 1:
                        if d not in pos_meaning:
                            pos_meaning.append(d)

            # 模式4: 短语翻译
            if not pos_meaning:
                ph_pattern = re.findall(
                    r'<div\s+class="[^"]*wordGroup[^"]*"[^>]*>'
                    r'.*?<span[^>]*>(.*?)</span>',
                    html, re.DOTALL
                )
                for p in ph_pattern:
                    p = re.sub(r'<[^>]+>', '', p).strip()
                    if p and len(p) > 1:
                        if p not in pos_meaning:
                            pos_meaning.append(p)

            if pos_meaning:
                return "; ".join(pos_meaning[:3])

            # fallback: 百度/必应风格结果
            fallback = re.findall(r'<div[^>]*class="[^"]*basic[^"]*"[^>]*>'
                                  r'.*?<li[^>]*>(.*?)</li>',
                                  html, re.DOTALL)
            if fallback:
                texts = []
                for f in fallback[:3]:
                    t = re.sub(r'<[^>]+>', '', f).strip()
                    if t:
                        texts.append(t)
                if texts:
                    return "; ".join(texts)

            return "未找到释义"

        except requests.RequestException:
            if attempt < retries - 1:
                time.sleep(1)
                continue
            return "查询失败"
    return "查询失败"


# ── 文本类型判断 ─────────────────────────────────

def is_scanned_pdf(pdf_bytes, sample_pages=2):
    """判断 PDF 是否为扫描件（无原生文本则视为扫描件）"""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    total = min(len(doc), sample_pages)
    text_len = 0
    for i in range(total):
        text_len += len(doc[i].get_text().strip())
    doc.close()
    return text_len < 50  # 几乎无原生文本即扫描件


# ── 主处理流程 ────────────────────────────────────

def process_pdf(pdf_bytes, filename="", on_status=None):
    """
    处理单个 PDF，返回单词数据列表：
    [{"word": str, "chinese": str, "sentence": str, "count": int}, ...]
    """
    if on_status:
        on_status(f"分析 {filename}...")

    scanned = is_scanned_pdf(pdf_bytes)
    if on_status:
        on_status(f"{'扫描件' if scanned else '文本'} PDF: {filename}")

    if scanned:
        raw = extract_scanned(pdf_bytes, filename)
    else:
        raw = extract_native(pdf_bytes, filename)

    # 收集所有黄色区域提取的文本
    word_context = {}  # lower(word) -> {"forms": {capitalized_form: count}, "sentences": []}

    for page_num in sorted(raw.keys()):
        page_data = raw[page_num]
        for item in page_data:
            text = item["text"]
            words = clean_text(text)
            words = [w for w in words if is_valid_word(w)]

            # 检测短语
            phrases = find_phrases(words)

            # 记录上下文（原文），按小写归一化
            for w in words:
                key = w.lower()
                if key not in word_context:
                    word_context[key] = {"forms": {}, "sentences": []}
                word_context[key]["forms"][w] = word_context[key]["forms"].get(w, 0) + 1
                context_snippet = text.strip()[:100]
                if context_snippet not in word_context[key]["sentences"]:
                    word_context[key]["sentences"].append(context_snippet)

            # 短语作为整体记录
            for phrase in phrases:
                if phrase not in word_context:
                    word_context[phrase] = {"forms": {phrase: 1}, "sentences": []}
                if text.strip()[:100] not in word_context[phrase]["sentences"]:
                    word_context[phrase]["sentences"].append(text.strip()[:100])

    if on_status:
        on_status(f"  共提取 {len(word_context)} 个单词/短语")

    # 查词典
    result = []
    total = len(word_context)
    for i, (key, data) in enumerate(sorted(word_context.items(),
                                            key=lambda x: -len(x[1]["sentences"]))):
        # 使用出现最多的形式作为显示单词
        best_form = max(data["forms"], key=data["forms"].get) if data["forms"] else key
        if on_status:
            on_status(f"  查词典 ({i+1}/{total}): {best_form}")
        chinese = query_youdao(key)
        sentence = data["sentences"][0] if data["sentences"] else ""
        result.append({
            "word": best_form,
            "chinese": chinese,
            "sentence": sentence,
            "count": len(data["sentences"]),
            "_key": key,  # 小写用于后续合并
        })
        time.sleep(0.3)  # 防止请求过快

    return result


def process_multiple_pdfs(file_dict, on_status=None):
    """
    处理多个 PDF。
    file_dict: {filename: bytes, ...}
    返回合并后的单词列表（按总出现次数降序）
    """
    # 收集所有结果
    all_results = {}  # word -> {chinese, sentences:[], total_count, sources:[]}

    total_files = len(file_dict)
    for idx, (fname, data) in enumerate(file_dict.items()):
        if on_status:
            on_status(f"[{idx+1}/{total_files}] {fname}")

        results = process_pdf(data, fname, on_status=on_status)

        for r in results:
            w = r["word"].lower()
            if w not in all_results:
                all_results[w] = {
                    "word": r["word"],
                    "chinese": r["chinese"],
                    "sentences": [],
                    "total_count": 0,
                    "sources": [],
                }
            if r["sentence"] and r["sentence"] not in all_results[w]["sentences"]:
                all_results[w]["sentences"].append(r["sentence"])
            all_results[w]["total_count"] += r["count"]
            if fname not in all_results[w]["sources"]:
                all_results[w]["sources"].append(fname)

    # 排序：总次数降序，同次数字母升序
    sorted_words = sorted(all_results.values(),
                          key=lambda x: (-x["total_count"], x["word"].lower()))

    return sorted_words


# ── Excel 生成 ────────────────────────────────────

def to_excel(words_data):
    """将单词数据导出为 Excel，返回 bytes"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "单词本"

    # 表头
    headers = ["单词/短语", "中文意思", "原文原句", "出现次数"]
    header_font = Font(bold=True, size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=11, name="Arial", color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 数据
    for i, item in enumerate(words_data, 2):
        ws.cell(row=i, column=1, value=item["word"])
        ws.cell(row=i, column=2, value=item["chinese"])
        # 兼容两种返回格式
        if "sentences" in item:
            sentences = "\n".join(item["sentences"][:3])
        else:
            sentences = item.get("sentence", "")
        ws.cell(row=i, column=3, value=sentences)
        ws.cell(row=i, column=4, value=item.get("total_count", item.get("count", 0)))

    # 列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12

    # 冻结首行
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
